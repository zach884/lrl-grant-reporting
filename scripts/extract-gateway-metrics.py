#!/usr/bin/env python3
"""Extract the seven Gateway semi-annual workbooks into reviewable JSON.

    python3 scripts/extract-gateway-metrics.py > reports/gateway-metrics-rows.json

WHY A SEPARATE STEP. Same reason as extract-sheet-rows.py: reading a funder's workbook and writing
activities to GHL are different concerns, only the second belongs in the app, and a JSON artifact
lets a person read what was parsed before anything reaches live.

WHY THESE WORKBOOKS AND NOT THE CONTACTS. A contact holds only its MOST RECENT Client Reporting
answers and records no submission date, so a contact-driven backfill can only file every snapshot
under one assumed period — which would fabricate history. These are seven funder-submitted
semi-annual reports already reconciled and sent to MEDC. See docs/sprints/gateway-metrics-import.md.

THE NOMINAL SUBMISSION DATE IS THE POINT. Each workbook is stamped with the date it was submitted
for, and `reportingPeriodFor()` turns that into the six-month window it describes. The period is half
the idempotency key, so it is derived here once, recorded in the JSON, and asserted by test — never
re-guessed downstream.
"""
import json, sys, re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")

ROOT = Path(__file__).resolve().parents[2] / "Past Grant Reports" / "Gateway"

# Filename → the date the workbook was submitted for. Gateway's April/October cadence lands exactly
# on the Feb-end/Aug-end boundaries `reportingPeriodFor()` already uses, so no new period logic is
# needed — but the mapping from FILE to date must be explicit. The filenames are inconsistent
# ("Apr2023", "Apr24", "Oct2025"), and inferring a year from two digits is how a 2023 workbook ends
# up filed under 2024.
WORKBOOKS = [
    ("Apr2023_LeanRocketLab_Company Metrics- NAICS (1).xlsx", "2023-04-15", "gateway-apr-2023"),
    ("Oct2023_LRL_Company Metrics- NAICS.xlsx",               "2023-10-15", "gateway-oct-2023"),
    ("Apr24_Lean_Rocket_Lab_Company Metrics- NAICS.xlsx",     "2024-04-15", "gateway-apr-2024"),
    ("Oct24_Lean_Rocket_Lab_Company Metrics- NAICS.xlsx",      "2024-10-15", "gateway-oct-2024"),
    ("Apr25_Lean_Rocket_Lab_Company Metrics- NAICS.xlsx",     "2025-04-15", "gateway-apr-2025"),
    ("Oct2025_Lean_Rocket_Lab_Company Metrics- NAICS.xlsx",   "2025-10-15", "gateway-oct-2025"),
    ("Apr26_Lean_Rocket_Lab_Company Metrics- NAICS (1).xlsx", "2026-04-15", "gateway-apr-2026"),
]

SHEET = "Companies Served"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

# Column letter → (expected header, output field). Verified 2026-09-02: all 15 headers are
# byte-identical across all seven workbooks, so the expected header is ASSERTED rather than
# prefix-matched. A funder silently re-ordering a column is exactly the failure a positional map
# cannot survive, and the assertion turns it into a loud error instead of shifted data.
COLUMNS = [
    ("C", "Company Name",                               "company_name"),
    ("K", "Email Address",                               "email"),
    ("N", "Commercialized Products",                     "products_commercialized"),
    ("O", "Products in the Commercialization Pipeline",  "products_in_pipeline"),
    ("P", "Jobs Created",                                "jobs_created"),
    ("Q", "Jobs Retained",                               "jobs_retained"),
    ("R", "MEDC Funds Awarded to Companies",             "medc_funding"),
    ("S", "SBIR, STTR, & Other Federal Funding",         "federal_funding"),
    ("T", "Venture Capital",                             "venture_capital"),
    ("U", "Angel Funds",                                 "angel_funding"),
    ("V", "Bank/Loan",                                   "bank_loans"),
    ("W", "Owner Investment",                            "owner_investment"),
    ("X", "New Sales (Increase in Revenue)",             "new_sales"),
    ("Y", "Other",                                       "other_funding"),
    ("Z", "Other Explanation",                           "other_explanation"),
]

NUMERIC = {
    "products_commercialized", "products_in_pipeline", "jobs_created", "jobs_retained",
    "medc_funding", "federal_funding", "venture_capital", "angel_funding", "bank_loans",
    "owner_investment", "new_sales", "other_funding",
}


def col_index(letter):
    n = 0
    for ch in letter:
        n = n * 26 + (ord(ch) - 64)
    return n


def norm(v):
    return re.sub(r"\s+", " ", str(v or "")).strip()


def number(v):
    """A funder's cell holds 0, '0', '', '-', '$1,200.00' or None. Only a real figure is a figure.

    Returns None for "not answered" and a float for anything numeric — INCLUDING zero, which is a
    reported zero and not a blank. Gateway's own instruction is to enter 0, so collapsing 0 to None
    would turn "we raised nothing" into "we did not report", and the two mean different things in a
    funder reconciliation.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "")
    if s in ("", "-", "--", "n/a", "N/A", "na", "NA", "TBD"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def main():
    out, problems = [], []
    for filename, submitted_at, slug in WORKBOOKS:
        path = ROOT / filename
        if not path.exists():
            problems.append({"workbook": filename, "problem": "file not found"})
            continue
        ws = openpyxl.load_workbook(path, data_only=True)[SHEET]

        for letter, expected, _field in COLUMNS:
            got = norm(ws.cell(HEADER_ROW, col_index(letter)).value)
            if got.lower() != expected.lower():
                problems.append({
                    "workbook": filename, "column": letter,
                    "problem": f"expected header {expected!r}, found {got!r}",
                })

        for r in range(FIRST_DATA_ROW, ws.max_row + 1):
            row = {"source_slug": slug, "workbook": filename, "submitted_at": submitted_at, "row": r}
            for letter, _expected, field in COLUMNS:
                raw = ws.cell(r, col_index(letter)).value
                row[field] = number(raw) if field in NUMERIC else (norm(raw) or None)
            # A "Companies Served" tab runs to row 400+ with only formatting below the data. A row
            # with neither a company nor an email is padding, not a company reporting nothing.
            if not row["company_name"] and not row["email"]:
                continue
            row["email"] = row["email"].lower() if row["email"] else None
            out.append(row)

    json.dump({
        "generatedAt": __import__("datetime").datetime.now(__import__("datetime").timezone.utc)
                        .isoformat(timespec="seconds").replace("+00:00", "Z"),
        "workbooks": [{"workbook": w, "submittedAt": d, "slug": s} for w, d, s in WORKBOOKS],
        "problems": problems,
        "rows": out,
    }, sys.stdout, indent=1)


if __name__ == "__main__":
    main()
