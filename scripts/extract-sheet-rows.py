#!/usr/bin/env python3
"""Extract the workflow-written grant spreadsheets into reviewable JSON.

    python3 scripts/extract-sheet-rows.py > reports/sheet-rows.json

WHY A SEPARATE STEP. Reading a funder's spreadsheet and creating activities in GHL are two different
concerns, and only the second one needs to be in the app. Keeping extraction here means the app takes
no xlsx dependency for what is a ONE-TIME historical import (ongoing capture is the webhooks plus the
nightly sweep), and it leaves a JSON artifact a human can read before anything is written to live.

The row NUMBER is carried through, because it is half of the idempotency key
(`tc-cumulative:row-<N>`). These sheets are appended by a GHL workflow and never re-sorted, so the
number is stable; a content hash was rejected because editing a note later must not orphan the
activity and create a second one.
"""
import json, sys, datetime, re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip3 install openpyxl")

ROOT = Path(__file__).resolve().parents[2] / "Past Grant Reports"

# Header labels carry embedded newlines and trailing spaces in both sheets, so a prefix match is
# needed. But EXACT wins first, because prefixes collide in ways that are silently wrong: "ST" (the
# state column) also prefixes "Street Address", and "Other" (the service flag) also prefixes "Other
# (Name)" and "Other Empty?". The first pass of this extractor put a street address in the state field
# for all 375 rows, which is the kind of error that survives review because the column is still full.
def find_col(header, wanted):
    flat = [str(h).replace("\n", " ").strip() if h else "" for h in header]
    for i, h in enumerate(flat):
        if h.lower() == wanted.lower():
            return i
    for i, h in enumerate(flat):
        if h and h.lower().startswith(wanted.lower()):
            return i
    return None

def cell(row, i):
    if i is None or i >= len(row):
        return None
    v = row[i]
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.date().isoformat() if isinstance(v, datetime.datetime) else v.isoformat()
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v

def zip_str(v):
    """Zips arrive as floats (49203.0). Keep the leading zeros a Michigan zip never has, but never
    emit '49203.0' into a funder's Zip Code column."""
    if v is None or v == "":
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip() or None

def truthy(v):
    return str(v).strip().lower() in ("true", "yes", "x", "1")

def extract(path, sheet, header_row, spec, source_slug):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    header = [r for r in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True)][0]
    cols = {name: find_col(header, prefix) for name, prefix in spec.items()}
    missing = [n for n, i in cols.items() if i is None]
    out = []
    for n, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        name = cell(row, cols["business_name"])
        if not name:
            continue
        out.append({
            "source_slug": source_slug,
            "row": n,
            "business_name": name,
            "owner_name": cell(row, cols.get("owner_name")),
            "email": cell(row, cols.get("email")),
            "date_added": cell(row, cols.get("date_added")),
            "county": cell(row, cols.get("county")),
            "address": cell(row, cols.get("address")),
            "city": cell(row, cols.get("city")),
            "state": cell(row, cols.get("state")),
            "zip": zip_str(cell(row, cols.get("zip"))),
            "notes": cell(row, cols.get("notes")),
            "flags": {k: truthy(row[cols[k]]) if cols.get(k) is not None else False
                      for k in spec if k.startswith("flag_")},
            "grant_amount": cell(row, cols.get("grant_amount")),
            "grant_date": cell(row, cols.get("grant_date")),
            "referral_reason": cell(row, cols.get("referral_reason")),
            # The referral TARGET, which lives in one of three named columns. These are what TC
            # columns V–AA actually report, and they are what makes two same-day referrals to
            # different partners distinguishable — without them, four separate referrals for one
            # company look like one row repeated four times.
            "referral_capital_provider": cell(row, cols.get("referral_capital_provider")),
            "referral_sb_partner": cell(row, cols.get("referral_sb_partner")),
            "referral_other": cell(row, cols.get("referral_other")),
            "referral_mentor": cell(row, cols.get("referral_mentor")),
            "referral_other_sbsh": cell(row, cols.get("referral_other_sbsh")),
            "referral_misbdc": cell(row, cols.get("referral_misbdc")),
            "referral_smartzone": cell(row, cols.get("referral_smartzone")),
        })
    return out, missing

TC = ("Trusted Connector Report.xlsx", "Cumulative Reporting", 1, {
    "date_added": "Date Added", "business_name": "Business Name", "address": "Street Address",
    "city": "City", "state": "ST", "zip": "Zip Code", "county": "County",
    "owner_name": "Business Owners", "email": "Email",
    "flag_one_on_one": "1:1 Technical Assistance", "flag_group": "Group Technical Assistance",
    "flag_event": "Hosted a Tech", "flag_networking": "Networking",
    "flag_referral": "Referral", "flag_other": "Other",
    "grant_amount": "Direct Grant", "grant_date": "Date Direct",
    "referral_reason": "Reason for Referral", "notes": "Notes",
    "referral_capital_provider": "Capital Provider Referral",
    "referral_sb_partner": "Small Buisness Ecosystem Partner Referral",
    "referral_other": "Other (Name)",
}, "tc-cumulative")

SBSH = ("SBSH Companies Served Spreadsheet (1).xlsx", "Sheet1", 1, {
    "date_added": "Date Added", "business_name": "Business Name", "address": "Street Address",
    "city": "City", "state": "ST", "zip": "Zip Code", "county": "County",
    "owner_name": "Business Owners", "email": "Email",
    "flag_one_on_one": "1:1 Business Consulting", "flag_group": "Group Training",
    "flag_support": "Small Business Support Services", "flag_other": "Other",
    "grant_amount": "Direct Grant", "grant_date": "Date Direct",
    "referral_reason": "Reason for Referral", "notes": "Notes",
    "referral_mentor": "Mentor Name", "referral_other_sbsh": "Other SBSH (Name)",
    "referral_misbdc": "MI-SBDC (Name)", "referral_smartzone": "SmartZone (Name)",
    "referral_other": "Other (Name)",
}, "sbsh-companies")

def main():
    result, report = [], []
    for fname, sheet, hdr, spec, slug in (TC, SBSH):
        path = ROOT / fname
        if not path.exists():
            report.append(f"SKIP {fname} (not found)")
            continue
        rows, missing = extract(path, sheet, hdr, spec, slug)
        result.extend(rows)
        report.append(f"{slug}: {len(rows)} rows" + (f"  ⚠️ columns not found: {missing}" if missing else ""))
    print(json.dumps({"generatedAt": datetime.datetime.now().isoformat(timespec="seconds"),
                      "rows": result}, indent=1))
    for line in report:
        print(line, file=sys.stderr)

if __name__ == "__main__":
    main()
